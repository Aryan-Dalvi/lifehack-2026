from __future__ import annotations

import csv
import io
import json
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from typing import Any

from fastapi import APIRouter, Request
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from app.db import connect, json_load, transaction, utc_now
from app.errors import api_error
from app.ids import new_id
from app.settings import settings

merchant_router = APIRouter(prefix="/merchant", tags=["merchant"])
catalog_router = APIRouter(prefix="/catalog", tags=["catalog"])
consumer_router = APIRouter(prefix="/consumer", tags=["consumer"])


class OnboardRequest(BaseModel):
    name: str = Field(min_length=1, max_length=100)
    size: str = "sme"
    category: str = "skincare"
    currency: str = "SGD"


class ConfigUpdate(BaseModel):
    name: str | None = None
    size: str | None = None
    currency: str | None = None
    accent_color: str | None = None
    persona: str | None = None
    policies: dict[str, Any] | None = None
    status: str | None = None


def _snippet(merchant_id: str) -> str:
    return (
        f'<script src="{settings.web_base_url}/widget.js" '
        f'data-merchant="{merchant_id}" data-position="bottom-right"></script>'
    )


def _merchant_payload(row) -> dict[str, Any]:
    return {
        "merchant_id": row["merchant_id"],
        "name": row["name"],
        "size": row["size"],
        "category": row["category"],
        "currency": row["currency"],
        "accent_color": row["accent_color"],
        "persona": row["persona"],
        "policies": json_load(row["policies_json"], {}),
        "status": row["status"],
        "hosted_url": f"{settings.web_base_url}/storefront?merchant={row['merchant_id']}",
        "embed_snippet": _snippet(row["merchant_id"]),
    }


@merchant_router.post("/onboard")
def onboard(body: OnboardRequest) -> dict[str, Any]:
    if body.category != "skincare":
        raise api_error(400, "VALIDATION", "The Phase 0 agent supports skincare only.")
    if body.size not in {"sme", "enterprise"}:
        raise api_error(400, "VALIDATION", "Merchant size must be sme or enterprise.")
    merchant_id = new_id("m")
    with transaction() as connection:
        connection.execute(
            "INSERT INTO merchants(merchant_id,name,size,category,currency,created_at) VALUES (?,?,?,?,?,?)",
            (merchant_id, body.name.strip(), body.size, "skincare", body.currency, utc_now()),
        )
    return {
        "merchant_id": merchant_id,
        "api_key": f"demo_{new_id('key')}",
        "embed_snippet": _snippet(merchant_id),
        "hosted_url": f"{settings.web_base_url}/storefront?merchant={merchant_id}",
    }


@merchant_router.get("/{merchant_id}/config")
def merchant_config(merchant_id: str) -> dict[str, Any]:
    with connect() as connection:
        row = connection.execute(
            "SELECT * FROM merchants WHERE merchant_id=?", (merchant_id,)
        ).fetchone()
    if not row:
        raise api_error(404, "NO_MERCHANT", "The merchant was not found.")
    return _merchant_payload(row)


@merchant_router.put("/{merchant_id}/config")
def update_config(merchant_id: str, body: ConfigUpdate) -> dict[str, Any]:
    values = body.model_dump(exclude_none=True)
    if values.get("size") not in {None, "sme", "enterprise"}:
        raise api_error(400, "VALIDATION", "Merchant size must be sme or enterprise.")
    if values.get("status") not in {None, "draft", "published"}:
        raise api_error(400, "VALIDATION", "Merchant status must be draft or published.")
    if not values:
        return merchant_config(merchant_id)

    columns: list[str] = []
    parameters: list[Any] = []
    for key, value in values.items():
        column = "policies_json" if key == "policies" else key
        columns.append(f"{column}=?")
        parameters.append(json.dumps(value) if key == "policies" else value)
    parameters.append(merchant_id)
    with transaction() as connection:
        updated = connection.execute(
            f"UPDATE merchants SET {', '.join(columns)} WHERE merchant_id=?", parameters
        ).rowcount
    if not updated:
        raise api_error(404, "NO_MERCHANT", "The merchant was not found.")
    return merchant_config(merchant_id)


@merchant_router.get("/{merchant_id}/snippet")
def merchant_snippet(merchant_id: str) -> dict[str, str]:
    merchant_config(merchant_id)
    return {
        "snippet": _snippet(merchant_id),
        "hosted_url": f"{settings.web_base_url}/storefront?merchant={merchant_id}",
    }


def _rows_from_csv(content: bytes) -> list[dict[str, Any]]:
    try:
        decoded = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise api_error(400, "BAD_CATALOG", "The CSV must use UTF-8 encoding.") from exc
    sample = decoded[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return list(csv.DictReader(io.StringIO(decoded), dialect=dialect))


def _rows_from_xlsx(content: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook.active
        values = list(worksheet.iter_rows(values_only=True))
    except Exception as exc:
        raise api_error(400, "BAD_CATALOG", "The XLSX file could not be read.") from exc
    if not values:
        return []
    headers = [str(value or "").strip() for value in values[0]]
    return [dict(zip(headers, row, strict=False)) for row in values[1:] if any(row)]


def _pick(row: dict[str, Any], *names: str) -> Any:
    normalized = {str(key).strip().lower().replace(" ", "_"): value for key, value in row.items()}
    for name in names:
        if name in normalized and normalized[name] not in {None, ""}:
            return normalized[name]
    return None


def _list_value(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip().lower() for item in value if str(item).strip()]
    return [item.strip().lower() for item in str(value).replace("|", ",").split(",") if item.strip()]


def _boolean(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    if value is None or value == "":
        return None
    normalized = str(value).strip().lower()
    if normalized in {"yes", "true", "1", "y"}:
        return True
    if normalized in {"no", "false", "0", "n"}:
        return False
    return None


def _price_cents(row: dict[str, Any]) -> int:
    cents = _pick(row, "price_cents")
    if cents is not None:
        return int(cents)
    price = _pick(row, "price", "price_sgd", "unit_price")
    if price is None:
        raise ValueError("price is required")
    normalized = str(price).replace("S$", "").replace("$", "").replace(",", "").strip()
    try:
        return int((Decimal(normalized) * 100).quantize(Decimal(1), rounding=ROUND_HALF_UP))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("price must be numeric") from exc


def _normalize_product(row: dict[str, Any], merchant: dict[str, Any]) -> dict[str, Any]:
    sku = str(_pick(row, "sku", "item_code", "product_id") or "").strip()
    title = str(_pick(row, "title", "name", "product_name") or "").strip()
    if not sku:
        raise ValueError("sku is required")
    if not title:
        raise ValueError("name is required")
    price_cents = _price_cents(row)
    if price_cents < 0:
        raise ValueError("price cannot be negative")
    stock = int(_pick(row, "stock", "inventory", "quantity") or 0)
    if stock < 0:
        raise ValueError("stock cannot be negative")
    ingredients = _list_value(_pick(row, "ingredients", "ingredient_list"))
    if not ingredients:
        raise ValueError("ingredients is required for skincare")
    attributes = {
        "routine_step": str(_pick(row, "routine_step", "product_type") or "treatment").lower(),
        "skin_types": _list_value(_pick(row, "skin_types", "skin_type")),
        "concerns": _list_value(_pick(row, "concerns", "skin_concerns")),
        "ingredients": ingredients,
        "excludes": _list_value(_pick(row, "excludes", "free_from")),
        "fragrance_free": _boolean(_pick(row, "fragrance_free")),
        "texture": str(_pick(row, "texture") or "").strip().lower(),
        "size_ml": int(_pick(row, "size_ml", "size") or 0),
    }
    rating_avg = _pick(row, "rating_avg", "rating")
    rating_count = _pick(row, "rating_count", "reviews")
    if rating_avg is not None:
        rating_avg = round(float(rating_avg), 1)
        if not 0 <= rating_avg <= 5:
            rating_avg = None
            rating_count = None
    return {
        "sku": sku,
        "merchant_id": merchant["merchant_id"],
        "title": title,
        "description": str(_pick(row, "description", "details") or "").strip(),
        "price_cents": price_cents,
        "currency": merchant["currency"],
        "image_url": str(_pick(row, "image_url", "image", "photo") or "").strip() or None,
        "category": "skincare",
        "attributes": attributes,
        "stock": stock,
        "rating_avg": rating_avg,
        "rating_count": int(rating_count) if rating_count is not None else None,
        "rating_source": "merchant_feed" if rating_avg is not None else "none",
    }


async def _catalog_rows(request: Request) -> tuple[list[dict[str, Any]], dict[str, str]]:
    content_type = request.headers.get("content-type", "")
    if "multipart/form-data" in content_type:
        form = await request.form()
        upload = form.get("file")
        if upload is None or not hasattr(upload, "read"):
            raise api_error(400, "BAD_CATALOG", "Choose a CSV, XLSX or JSON catalog file.")
        content = await upload.read()
        if len(content) > 5 * 1024 * 1024:
            raise api_error(413, "TOO_LARGE", "Catalog files are limited to 5 MB.")
        filename = getattr(upload, "filename", "catalog.csv") or "catalog.csv"
        extension = filename.lower().rsplit(".", 1)[-1]
        if extension == "csv":
            rows = _rows_from_csv(content)
        elif extension == "xlsx":
            rows = _rows_from_xlsx(content)
        elif extension == "json":
            parsed = json.loads(content)
            rows = parsed.get("products", parsed) if isinstance(parsed, dict) else parsed
        else:
            raise api_error(400, "BAD_CATALOG", "Only CSV, XLSX and JSON catalogs are supported.")
        return rows, {"filename": filename, "format": extension}

    if "application/json" in content_type:
        parsed = await request.json()
        rows = parsed.get("products", []) if isinstance(parsed, dict) else parsed
        return rows, {"filename": "catalog.json", "format": "json"}
    raise api_error(400, "BAD_CATALOG", "Use multipart file upload or application/json.")


@merchant_router.post("/{merchant_id}/catalog")
async def ingest_catalog(merchant_id: str, request: Request) -> dict[str, Any]:
    merchant = merchant_config(merchant_id)
    rows, source = await _catalog_rows(request)
    if not isinstance(rows, list) or not rows:
        raise api_error(400, "BAD_CATALOG", "The catalog contains no product rows.")
    normalized: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row_number, row in enumerate(rows, start=2):
        try:
            product = _normalize_product(row, merchant)
            if product["sku"] in seen:
                raise ValueError(f"duplicate sku {product['sku']}; kept first")
            seen.add(product["sku"])
            normalized.append(product)
        except (TypeError, ValueError, InvalidOperation) as exc:
            errors.append({"row": row_number, "reason": str(exc)})

    now = utc_now()
    with transaction() as connection:
        for product in normalized:
            connection.execute(
                "INSERT INTO products(sku,merchant_id,title,description,price_cents,currency,image_url,"
                "category,attributes_json,stock,rating_avg,rating_count,rating_source,created_at,updated_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(sku) DO UPDATE SET "
                "title=excluded.title,description=excluded.description,price_cents=excluded.price_cents,"
                "image_url=excluded.image_url,attributes_json=excluded.attributes_json,stock=excluded.stock,"
                "rating_avg=excluded.rating_avg,rating_count=excluded.rating_count,"
                "rating_source=excluded.rating_source,updated_at=excluded.updated_at",
                (
                    product["sku"],
                    product["merchant_id"],
                    product["title"],
                    product["description"],
                    product["price_cents"],
                    product["currency"],
                    product["image_url"],
                    product["category"],
                    json.dumps(product["attributes"]),
                    product["stock"],
                    product["rating_avg"],
                    product["rating_count"],
                    product["rating_source"],
                    now,
                    now,
                ),
            )
    mappings = {
        "sku": "SKU",
        "title": "Name",
        "price_cents": "Price",
        "ingredients": "Ingredients",
        "skin_types": "Skin types",
        "stock": "Stock",
    }
    return {
        "ingested": len(normalized),
        "skipped": len(errors),
        "errors": errors,
        "source": source,
        "mappings": mappings,
        "partial_success": bool(normalized and errors),
    }


def _product_payload(row) -> dict[str, Any]:
    return {
        "sku": row["sku"],
        "merchant_id": row["merchant_id"],
        "merchant_name": row["merchant_name"],
        "merchant_size": row["merchant_size"],
        "title": row["title"],
        "description": row["description"],
        "price_cents": int(row["price_cents"]),
        "currency": row["currency"],
        "image_url": row["image_url"],
        "category": row["category"],
        "attributes": json_load(row["attributes_json"], {}),
        "stock": int(row["stock"]),
        "rating_avg": row["rating_avg"],
        "rating_count": row["rating_count"],
        "rating_source": row["rating_source"],
    }


@catalog_router.get("/search")
def catalog_search(
    q: str = "",
    merchant_id: str = "m_mysa",
    category: str = "skincare",
    max_price_cents: int | None = None,
    attrs: str | None = None,
    limit: int = 5,
) -> dict[str, Any]:
    if category != "skincare":
        raise api_error(400, "VALIDATION", "The Phase 0 catalog supports skincare only.")
    limit = max(1, min(limit, 5))
    try:
        attr_filters = json.loads(attrs) if attrs else {}
    except json.JSONDecodeError as exc:
        raise api_error(400, "VALIDATION", "Attribute filters must be valid JSON.") from exc
    with connect() as connection:
        rows = connection.execute(
            "SELECT p.*,m.name AS merchant_name,m.size AS merchant_size FROM products p "
            "JOIN merchants m ON m.merchant_id=p.merchant_id "
            "WHERE p.merchant_id=? AND p.category='skincare' AND p.stock>0",
            (merchant_id,),
        ).fetchall()
    query_tokens = {token for token in q.lower().replace("-", " ").split() if len(token) > 2}
    candidates: list[tuple[int, dict[str, Any]]] = []
    for row in rows:
        product = _product_payload(row)
        if max_price_cents is not None and product["price_cents"] > max_price_cents:
            continue
        attributes = product["attributes"]
        if any(
            isinstance(expected, list)
            and not set(map(str.lower, expected)).issubset(set(map(str.lower, attributes.get(key, []))))
            for key, expected in attr_filters.items()
            if isinstance(expected, list)
        ):
            continue
        rejected = False
        for key, expected in attr_filters.items():
            actual = attributes.get(key)
            if isinstance(expected, list):
                continue
            if isinstance(expected, bool) and actual is not expected or expected is not None and not isinstance(expected, bool) and str(actual).lower() != str(expected).lower():
                rejected = True
        if rejected:
            continue
        searchable = " ".join(
            [
                product["title"],
                product["description"],
                json.dumps(attributes),
            ]
        ).lower()
        score = sum(4 if token in product["title"].lower() else 1 for token in query_tokens if token in searchable)
        if query_tokens and score == 0:
            continue
        candidates.append((score, product))
    candidates.sort(
        key=lambda entry: (
            -entry[0],
            -(entry[1]["rating_avg"] or 0),
            entry[1]["price_cents"],
            entry[1]["sku"],
        )
    )
    results = [product for _, product in candidates[:limit]]
    return {
        "results": results,
        "total": len(candidates),
        "facets": {
            "routine_steps": sorted(
                {product["attributes"].get("routine_step") for _, product in candidates}
                - {None, ""}
            ),
            "currency": "SGD",
        },
        "source": "catalog_database",
    }


@catalog_router.get("/product/{sku}")
def catalog_product(sku: str) -> dict[str, Any]:
    with connect() as connection:
        row = connection.execute(
            "SELECT p.*,m.name AS merchant_name,m.size AS merchant_size FROM products p "
            "JOIN merchants m ON m.merchant_id=p.merchant_id WHERE p.sku=?",
            (sku,),
        ).fetchone()
    if not row:
        raise api_error(404, "NO_PRODUCT", "The product was not found.")
    return _product_payload(row)


@consumer_router.get("/{consumer_id}/addresses")
def addresses(consumer_id: str) -> dict[str, Any]:
    with connect() as connection:
        rows = connection.execute(
            "SELECT * FROM addresses WHERE consumer_id=? ORDER BY is_default DESC", (consumer_id,)
        ).fetchall()
    if not rows:
        raise api_error(404, "NO_CONSUMER", "No saved addresses were found.")
    values = [
        {
            "address_id": row["address_id"],
            "consumer_id": row["consumer_id"],
            "recipient": row["recipient"],
            "lines": json_load(row["lines_json"], []),
            "postal_code": row["postal_code"],
            "country": row["country"],
            "is_default": bool(row["is_default"]),
        }
        for row in rows
    ]
    return {
        "addresses": values,
        "default_address_id": next(
            (value["address_id"] for value in values if value["is_default"]), None
        ),
    }


@consumer_router.put("/{consumer_id}/addresses/{address_id}/default")
def set_default_address(consumer_id: str, address_id: str) -> dict[str, str]:
    with transaction() as connection:
        exists = connection.execute(
            "SELECT 1 FROM addresses WHERE consumer_id=? AND address_id=?",
            (consumer_id, address_id),
        ).fetchone()
        if not exists:
            raise api_error(404, "NO_ADDRESS", "The shipping address was not found.")
        connection.execute("UPDATE addresses SET is_default=0 WHERE consumer_id=?", (consumer_id,))
        connection.execute(
            "UPDATE addresses SET is_default=1 WHERE address_id=?", (address_id,)
        )
    return {"default_address_id": address_id}

