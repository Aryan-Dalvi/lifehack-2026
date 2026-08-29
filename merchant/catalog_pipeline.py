from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

from app.db import connect, json_load, transaction, utc_now
from app.errors import api_error
from app.ids import new_id
from app.settings import settings
from merchant.catalog_cleaner import (
    CLEANER_VERSION,
    PROMPT_HASH,
    TAXONOMY_VERSION,
    assignments_by_axis,
    build_taxonomy,
    classify_records,
    normalize_key,
)
from merchant.catalog_diagnostics import summarize_upload
from merchant.catalog_images import (
    ImageArchiveError,
    ImageTarget,
    extract_image_archive,
    match_images,
)
from merchant.catalog_mapping import (
    FIELD_ALIASES,
    MODEL_DESCRIPTIVE_FIELDS,
    MODEL_EXCLUDED_TARGETS,
    MappingCandidates,
    MappingResolution,
    detect_mappings,
    resolve_mappings,
)

CATALOG_SCHEMA_VERSION = "catalog-cleaning.v1"
JSON_SOURCE_SCHEMA_VERSION = "catalog-source.v1"
PARSER_VERSION = "catalog-parser.v1"
MAX_CATALOG_BYTES = 5 * 1024 * 1024
MAX_EXPANDED_XLSX_BYTES = 50 * 1024 * 1024
MAX_CATALOG_ROWS = 10_000
MAX_CATALOG_COLUMNS = 200
MAX_PREVIEW_ROWS = 100

@dataclass(frozen=True)
class SourceRow:
    source_record_id: str
    row_number: int
    sheet_name: str | None
    values: dict[str, Any]


@dataclass(frozen=True)
class ParsedCatalog:
    filename: str
    source_format: str
    raw_bytes: bytes
    rows: list[SourceRow]
    metadata: dict[str, Any]
    candidates: MappingCandidates

    @property
    def mappings(self) -> dict[str, str]:
        """The unambiguous alias matches. Ties are settled by resolve_mappings."""
        return self.candidates.resolved

    @property
    def source_sha256(self) -> str:
        return hashlib.sha256(self.raw_bytes).hexdigest()


class ProductFactError(ValueError):
    def __init__(self, issues: list[str]):
        super().__init__("; ".join(issues))
        self.issues = issues


class DuplicateJsonKeyError(ValueError):
    pass


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, dict):
        return {str(key): _json_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_value(item) for item in value]
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def canonical_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _catalog_snapshot(connection: Any, merchant_id: str) -> dict[str, Any]:
    rows = connection.execute(
        "SELECT sku,merchant_id,title,description,price_cents,currency,image_url,category,"
        "attributes_json,stock,rating_avg,rating_count,rating_source,created_at,updated_at "
        "FROM products WHERE merchant_id=? ORDER BY sku",
        (merchant_id,),
    ).fetchall()
    products = [dict(row) for row in rows]
    return {
        "hash": hashlib.sha256(canonical_json(products).encode()).hexdigest(),
        "skus": [row["sku"] for row in products],
    }


def _row_hash(values: dict[str, Any]) -> str:
    return hashlib.sha256(canonical_json(values).encode()).hexdigest()


def _validate_headers(headers: list[Any]) -> list[str]:
    normalized_headers = [str(value or "").strip() for value in headers]
    while normalized_headers and not normalized_headers[-1]:
        normalized_headers.pop()
    if not normalized_headers:
        raise api_error(400, "BAD_CATALOG", "The catalog has no header row.")
    if len(normalized_headers) > MAX_CATALOG_COLUMNS:
        raise api_error(
            400, "BAD_CATALOG", f"Catalogs are limited to {MAX_CATALOG_COLUMNS} columns."
        )
    if any(not header for header in normalized_headers):
        raise api_error(
            400, "BAD_CATALOG", "Catalog headers cannot be blank between named columns."
        )
    normalized_keys = [normalize_key(header) for header in normalized_headers]
    duplicates = sorted({key for key in normalized_keys if normalized_keys.count(key) > 1})
    if duplicates:
        raise api_error(
            400,
            "BAD_CATALOG",
            f"Catalog headers must be unique; duplicate header: {duplicates[0]}.",
        )
    return normalized_headers


def _parse_csv(content: bytes, filename: str) -> ParsedCatalog:
    if content.startswith((b"\xff\xfe", b"\xfe\xff")):
        encoding = "utf-16"
        decoded = content.decode(encoding)
    else:
        try:
            encoding = "utf-8-sig"
            decoded = content.decode(encoding)
        except UnicodeDecodeError:
            try:
                encoding = "cp1252"
                decoded = content.decode(encoding)
            except UnicodeDecodeError as exc:
                raise api_error(
                    400,
                    "BAD_CATALOG",
                    "The CSV encoding could not be detected; use UTF-8, UTF-16, or Windows-1252.",
                ) from exc
    sample = decoded[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(decoded, newline=""), dialect=dialect)
    try:
        headers = _validate_headers(next(reader))
    except StopIteration as exc:
        raise api_error(400, "BAD_CATALOG", "The CSV is empty.") from exc
    rows: list[SourceRow] = []
    for row_number, values in enumerate(reader, start=2):
        if not any(value not in (None, "") for value in values):
            continue
        if len(values) > len(headers) and any(
            value not in (None, "") for value in values[len(headers) :]
        ):
            raise api_error(
                400,
                "BAD_CATALOG",
                f"Row {row_number} contains more values than the header row.",
            )
        raw_row = {
            header: _json_value(values[index] if index < len(values) else None)
            for index, header in enumerate(headers)
        }
        rows.append(
            SourceRow(
                source_record_id=f"row_{row_number:06d}",
                row_number=row_number,
                sheet_name=None,
                values=raw_row,
            )
        )
        if len(rows) > MAX_CATALOG_ROWS:
            raise api_error(413, "TOO_LARGE", f"Catalogs are limited to {MAX_CATALOG_ROWS} rows.")
    parsed = ParsedCatalog(
        filename=filename,
        source_format="csv",
        raw_bytes=content,
        rows=rows,
        metadata={"delimiter": dialect.delimiter, "encoding": encoding},
        candidates=detect_mappings(rows),
    )
    return parsed


def _inspect_xlsx_archive(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            members = archive.infolist()
            expanded_size = sum(member.file_size for member in members)
            names = {member.filename.lower() for member in members}
    except (zipfile.BadZipFile, OSError) as exc:
        raise api_error(400, "BAD_CATALOG", "The XLSX file is not a valid workbook.") from exc
    if expanded_size > MAX_EXPANDED_XLSX_BYTES:
        raise api_error(413, "TOO_LARGE", "The expanded XLSX workbook is too large.")
    if any("vbaproject.bin" in name or "externallinks/" in name for name in names):
        raise api_error(
            400,
            "BAD_CATALOG",
            "Macro-enabled or externally linked workbooks are not accepted.",
        )


def _xlsx_cell_value(cell: Any) -> Any:
    value = cell.value
    if (
        isinstance(value, int)
        and isinstance(cell.number_format, str)
        and re.fullmatch(r"0+", cell.number_format)
    ):
        return str(value).zfill(len(cell.number_format))
    return _json_value(value)


def _worksheet_header_score(worksheet: Any) -> int:
    headers = [
        normalize_key(_xlsx_cell_value(cell))
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
        if _xlsx_cell_value(cell) not in (None, "")
    ]
    return sum(
        any(alias in headers for alias in aliases)
        for target, aliases in FIELD_ALIASES.items()
        if target in {"sku", "title", "price", "price_cents", "stock", "ingredients"}
    )


def _select_worksheet(
    workbook: Any, requested_sheet: str | None
) -> tuple[Any, list[dict[str, Any]]]:
    candidates = [
        {
            "name": worksheet.title,
            "state": worksheet.sheet_state,
            "header_score": _worksheet_header_score(worksheet),
        }
        for worksheet in workbook.worksheets
    ]
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise api_error(400, "BAD_CATALOG", f"Worksheet {requested_sheet!r} was not found.")
        return workbook[requested_sheet], candidates
    best_score = max(candidate["header_score"] for candidate in candidates)
    best_names = [
        candidate["name"] for candidate in candidates if candidate["header_score"] == best_score
    ]
    if best_score == 0:
        return workbook.active, candidates
    if len(best_names) > 1:
        raise api_error(
            400,
            "BAD_CATALOG",
            "Multiple worksheets look like product catalogs; choose a worksheet explicitly.",
            worksheet_names=best_names,
        )
    return workbook[best_names[0]], candidates


def _parse_xlsx(content: bytes, filename: str, requested_sheet: str | None = None) -> ParsedCatalog:
    _inspect_xlsx_archive(content)
    try:
        workbook = load_workbook(
            io.BytesIO(content), read_only=False, data_only=False, keep_links=False
        )
        worksheet, worksheet_candidates = _select_worksheet(workbook, requested_sheet)
        if worksheet.max_row < 1:
            raise api_error(400, "BAD_CATALOG", "The XLSX worksheet is empty.")
        if worksheet.max_column > MAX_CATALOG_COLUMNS:
            raise api_error(
                400,
                "BAD_CATALOG",
                f"Catalogs are limited to {MAX_CATALOG_COLUMNS} columns.",
            )
        if worksheet.max_row - 1 > MAX_CATALOG_ROWS:
            raise api_error(
                413,
                "TOO_LARGE",
                f"Catalogs are limited to {MAX_CATALOG_ROWS} rows.",
            )
        worksheet_name = worksheet.title
        worksheet_count = len(workbook.sheetnames)
        header_cells = list(next(worksheet.iter_rows(min_row=1, max_row=1)))
        headers = _validate_headers([_xlsx_cell_value(cell) for cell in header_cells])
        rows: list[SourceRow] = []
        formula_cells = 0
        for row_number, cells in enumerate(
            worksheet.iter_rows(min_row=2, max_col=len(headers)), start=2
        ):
            values = [_xlsx_cell_value(cell) for cell in cells]
            if not any(value not in (None, "") for value in values):
                continue
            formula_cells += sum(
                1
                for cell, value in zip(cells, values, strict=False)
                if cell.data_type == "f" or str(value).lstrip().startswith("=")
            )
            raw_row = {header: values[index] for index, header in enumerate(headers)}
            rows.append(
                SourceRow(
                    source_record_id=f"row_{row_number:06d}",
                    row_number=row_number,
                    sheet_name=worksheet.title,
                    values=raw_row,
                )
            )
            if len(rows) > MAX_CATALOG_ROWS:
                raise api_error(
                    413, "TOO_LARGE", f"Catalogs are limited to {MAX_CATALOG_ROWS} rows."
                )
    except Exception as exc:
        if getattr(exc, "status_code", None):
            raise
        raise api_error(400, "BAD_CATALOG", "The XLSX file could not be read.") from exc
    finally:
        if "workbook" in locals():
            workbook.close()
    parsed = ParsedCatalog(
        filename=filename,
        source_format="xlsx",
        raw_bytes=content,
        rows=rows,
        metadata={
            "worksheet": worksheet_name,
            "worksheet_count": worksheet_count,
            "worksheet_selection": "explicit" if requested_sheet else "header_match",
            "worksheet_candidates": worksheet_candidates,
            "formula_cells": formula_cells,
        },
        candidates=detect_mappings(rows),
    )
    return parsed


def _parse_json(content: bytes, filename: str) -> ParsedCatalog:
    def unique_object(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
        result: dict[str, Any] = {}
        normalized_seen: set[str] = set()
        for key, value in pairs:
            normalized = normalize_key(key)
            if normalized in normalized_seen:
                raise DuplicateJsonKeyError(f"duplicate JSON key {key!r}")
            normalized_seen.add(normalized)
            result[key] = value
        return result

    try:
        parsed = json.loads(content.decode("utf-8-sig"), object_pairs_hook=unique_object)
    except (DuplicateJsonKeyError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise api_error(400, "BAD_CATALOG", "The JSON catalog is malformed or not UTF-8.") from exc
    source_schema_version = parsed.get("schema_version") if isinstance(parsed, dict) else None
    if source_schema_version not in {None, JSON_SOURCE_SCHEMA_VERSION}:
        raise api_error(
            400,
            "BAD_CATALOG",
            f"Unsupported JSON schema_version at $.schema_version: {source_schema_version!r}.",
        )
    products = parsed.get("products", parsed) if isinstance(parsed, dict) else parsed
    if not isinstance(products, list):
        raise api_error(400, "BAD_CATALOG", "JSON catalogs must be an array or contain products[].")
    if len(products) > MAX_CATALOG_ROWS:
        raise api_error(413, "TOO_LARGE", f"Catalogs are limited to {MAX_CATALOG_ROWS} rows.")
    rows: list[SourceRow] = []
    for index, item in enumerate(products, start=1):
        if not isinstance(item, dict):
            raise api_error(400, "BAD_CATALOG", f"JSON products[{index - 1}] must be an object.")
        raw_row = {str(key): _json_value(value) for key, value in item.items()}
        if not any(value not in (None, "", [], {}) for value in raw_row.values()):
            continue
        rows.append(
            SourceRow(
                source_record_id=f"row_{index:06d}",
                row_number=index,
                sheet_name=None,
                values=raw_row,
            )
        )
    parsed_catalog = ParsedCatalog(
        filename=filename,
        source_format="json",
        raw_bytes=content,
        rows=rows,
        metadata={"schema_version": source_schema_version or "legacy-unversioned"},
        candidates=detect_mappings(rows),
    )
    return parsed_catalog


def parse_catalog(content: bytes, filename: str, *, sheet_name: str | None = None) -> ParsedCatalog:
    if len(content) > MAX_CATALOG_BYTES:
        raise api_error(413, "TOO_LARGE", "Catalog files are limited to 5 MB.")
    extension = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if extension == "csv":
        parsed = _parse_csv(content, filename)
    elif extension == "xlsx":
        parsed = _parse_xlsx(content, filename, requested_sheet=sheet_name)
    elif extension == "json":
        parsed = _parse_json(content, filename)
    else:
        raise api_error(400, "BAD_CATALOG", "Only CSV, XLSX and JSON catalogs are supported.")
    if not parsed.rows:
        raise api_error(400, "BAD_CATALOG", "The catalog contains no product rows.")
    return parsed


def _mapped(row: SourceRow, mappings: dict[str, str], target: str) -> Any:
    source = mappings.get(target)
    if not source:
        return None
    if source in row.values:
        return row.values[source]
    normalized_source = normalize_key(source)
    return next(
        (
            value
            for column, value in row.values.items()
            if normalize_key(column) == normalized_source
        ),
        None,
    )


def _clean_text(value: Any) -> str:
    return " ".join(unicodedata.normalize("NFKC", str(value or "")).split())


def _integer(value: Any, *, field: str, required: bool = False) -> int | None:
    if value in (None, ""):
        if required:
            raise ValueError(f"{field} is required")
        return None
    if isinstance(value, bool):
        raise TypeError(f"{field} must be an integer")
    try:
        decimal = Decimal(str(value).strip())
    except InvalidOperation as exc:
        raise ValueError(f"{field} must be an integer") from exc
    if not decimal.is_finite():
        raise ValueError(f"{field} must be a finite integer")
    if decimal != decimal.to_integral_value():
        raise ValueError(f"{field} must be an integer")
    return int(decimal)


def _price_cents(row: SourceRow, mappings: dict[str, str]) -> int:
    cents_value = _mapped(row, mappings, "price_cents")
    if cents_value not in (None, ""):
        cents = _integer(cents_value, field="price_cents", required=True)
        assert cents is not None
        return cents
    value = _mapped(row, mappings, "price")
    if value in (None, ""):
        raise ValueError("price is required")
    if isinstance(value, bool):
        raise TypeError("price must be numeric")
    if isinstance(value, (int, float, Decimal)):
        decimal = Decimal(str(value))
    else:
        normalized = (
            str(value).strip().replace("S$", "").replace("SGD", "").replace("$", "").strip()
        )
        if "," in normalized and "." not in normalized:
            if re.fullmatch(r"\d{1,3}(?:,\d{3})+", normalized):
                normalized = normalized.replace(",", "")
            else:
                raise ValueError("price uses an ambiguous comma decimal or thousands format")
        elif "," in normalized:
            if not re.fullmatch(r"\d{1,3}(?:,\d{3})+\.\d{1,2}", normalized):
                raise ValueError("price format is ambiguous")
            normalized = normalized.replace(",", "")
        if not re.fullmatch(r"\d+(?:\.\d{1,2})?", normalized):
            raise ValueError("price must be numeric with at most two decimal places")
        decimal = Decimal(normalized)
    if not decimal.is_finite():
        raise ValueError("price must be finite")
    if decimal.as_tuple().exponent < -2:
        raise ValueError("price must have at most two decimal places")
    try:
        return int((decimal * 100).quantize(Decimal(1), rounding=ROUND_HALF_UP))
    except InvalidOperation as exc:
        raise ValueError("price is outside the supported range") from exc


def _split_list(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    items = value if isinstance(value, list) else re.split(r"[,;|]+", str(value))
    result: list[str] = []
    for item in items:
        normalized = _clean_text(item).lower()
        if normalized and normalized not in result:
            result.append(normalized)
    return result


_LOCAL_IMAGE_HOSTS = ("localhost", "127.0.0.1")


def _image_url(value: Any) -> str | None:
    """A product image link a browser can safely load.

    An image URL is merchant-supplied text that ends up in an ``<img src>`` on every
    shopper's page, so only two shapes are accepted: an https address, or a path on this
    site. Everything else - javascript:, data:, protocol-relative //evil.example - is
    refused with a message the merchant can act on rather than being passed through.
    """
    text = _clean_text(value)
    if not text:
        return None
    if text.startswith("/") and not text.startswith("//"):
        return text
    lowered = text.lower()
    if lowered.startswith("https://"):
        return text
    if lowered.startswith("http://") and any(
        lowered.startswith(f"http://{host}") for host in _LOCAL_IMAGE_HOSTS
    ):
        return text
    raise ValueError(
        "image_url must be an https:// link or a path on this store, "
        "or be left blank and supplied in the image ZIP"
    )


def _boolean(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    if value in (None, ""):
        return None
    normalized = _clean_text(value).lower()
    if normalized in {"yes", "true", "1", "y"}:
        return True
    if normalized in {"no", "false", "0", "n"}:
        return False
    raise ValueError("fragrance_free must be true or false")


def normalize_locked_facts(
    row: SourceRow, merchant: dict[str, Any], mappings: dict[str, str]
) -> dict[str, Any]:
    issues: list[str] = []
    sku = _clean_text(_mapped(row, mappings, "sku"))
    title = _clean_text(_mapped(row, mappings, "title"))
    if not sku:
        issues.append("sku is required")
    if not title:
        issues.append("name is required")
    try:
        price_cents = _price_cents(row, mappings)
        if price_cents < 0:
            issues.append("price cannot be negative")
    except (TypeError, ValueError) as exc:
        price_cents = None
        issues.append(str(exc))
    try:
        stock = _integer(_mapped(row, mappings, "stock"), field="stock", required=True)
        if stock is not None and stock < 0:
            issues.append("stock cannot be negative")
    except (TypeError, ValueError) as exc:
        stock = None
        issues.append(str(exc))
    ingredients = _split_list(_mapped(row, mappings, "ingredients"))
    if not ingredients:
        issues.append("ingredients is required for skincare")

    rating_avg: float | None = None
    rating_count: int | None = None
    raw_rating = _mapped(row, mappings, "rating_avg")
    raw_rating_count = _mapped(row, mappings, "rating_count")
    if raw_rating not in (None, ""):
        try:
            rating_avg = round(float(raw_rating), 1)
            if not 0 <= rating_avg <= 5:
                raise ValueError
        except (TypeError, ValueError):
            issues.append("rating must be between 0 and 5")
            rating_avg = None
        try:
            rating_count = _integer(raw_rating_count, field="rating_count", required=True)
            if rating_count is not None and rating_count < 0:
                issues.append("rating_count cannot be negative")
        except (TypeError, ValueError) as exc:
            rating_count = None
            issues.append(str(exc))
    elif raw_rating_count not in (None, ""):
        issues.append("rating_count requires rating_avg")

    try:
        size_ml = _integer(_mapped(row, mappings, "size_ml"), field="size_ml")
        if size_ml is not None and size_ml <= 0:
            issues.append("size_ml must be positive")
    except (TypeError, ValueError) as exc:
        size_ml = None
        issues.append(str(exc))
    try:
        fragrance_free = _boolean(_mapped(row, mappings, "fragrance_free"))
    except ValueError as exc:
        fragrance_free = None
        issues.append(str(exc))

    try:
        image_url = _image_url(_mapped(row, mappings, "image_url"))
    except ValueError as exc:
        image_url = None
        issues.append(str(exc))

    source_currency = _clean_text(_mapped(row, mappings, "currency")).upper()
    merchant_currency = str(merchant["currency"]).upper()
    if source_currency and source_currency != merchant_currency:
        issues.append(
            f"source currency {source_currency} conflicts with merchant currency {merchant_currency}"
        )
    formula_fields = [
        column
        for column, value in row.values.items()
        if isinstance(value, str) and value.lstrip().startswith("=")
    ]
    if formula_fields:
        issues.append(f"formula content requires review: {formula_fields[0]}")
    if issues:
        raise ProductFactError(issues)
    assert price_cents is not None and stock is not None
    return {
        "sku": sku,
        "merchant_id": merchant["merchant_id"],
        "title": title,
        "description": _clean_text(_mapped(row, mappings, "description")),
        "price_cents": price_cents,
        "currency": merchant_currency,
        "image_url": image_url,
        "category": "skincare",
        "stock": stock,
        "ingredients": ingredients,
        "excludes": _split_list(_mapped(row, mappings, "excludes")),
        "fragrance_free": fragrance_free,
        "texture": _clean_text(_mapped(row, mappings, "texture")).lower(),
        "size_ml": size_ml,
        "rating_avg": rating_avg,
        "rating_count": rating_count,
        "rating_source": "merchant_feed" if rating_avg is not None else "none",
    }


def _descriptive_fields(row: SourceRow, mappings: dict[str, str]) -> dict[str, Any]:
    """Evidence the screening model may read.

    Two ways in, and a locked-fact column is excluded from both:

    1. a field the mapping step resolved to a descriptive target - this is what lets a
       header like `full_inci` still be read as the ingredient list, and
    2. a header whose own name is on the allow-list.

    (1) matters more than it looks: without it, evidence reached the model only when a
    merchant's header happened to be spelled the way the allow-list spells it, so a correctly
    mapped but differently named title column was screened with nothing to go on.
    """
    excluded_columns = {mappings[target] for target in MODEL_EXCLUDED_TARGETS if target in mappings}
    fields: dict[str, Any] = {}

    for target, column in mappings.items():
        if target in MODEL_EXCLUDED_TARGETS:
            continue
        value = row.values.get(column)
        if value not in (None, ""):
            fields[target] = value

    for column, value in row.values.items():
        if column in excluded_columns or column in mappings.values() or value in (None, ""):
            continue
        if normalize_key(column) in MODEL_DESCRIPTIVE_FIELDS:
            fields[column] = value
    return fields


def _canonical_product(
    *,
    facts: dict[str, Any],
    classification: dict[str, Any],
    upload_id: str,
    run_id: str,
    source_row: SourceRow,
    source_sha256: str,
) -> dict[str, Any]:
    axes = assignments_by_axis(classification)
    primary_types = [
        assignment["proposed_slug"]
        for assignment in classification["assignments"]
        if assignment["axis"] == "product_type" and assignment["is_primary"]
    ]
    product_type = primary_types[0] if len(primary_types) == 1 else None
    routine_steps = axes["routine_step"]
    attributes = {
        "product_type": product_type,
        "product_types": axes["product_type"],
        "routine_step": routine_steps[0] if routine_steps else None,
        "routine_steps": routine_steps,
        "skin_types": axes["skin_type"],
        "concerns": [slug.replace("_", " ") for slug in axes["concern"]],
        "ingredients": facts["ingredients"],
        "ingredient_entities": axes["ingredient_entity"],
        "ingredient_functions": axes["function"],
        "usage_times": axes["usage_time"],
        "formulations": axes["formulation"],
        "excludes": facts["excludes"],
        "free_from": axes["free_from"],
        "fragrance_free": facts["fragrance_free"],
        "texture": facts["texture"],
        "size_ml": facts["size_ml"],
        "categories": [
            f"{axis}:{slug}"
            for axis, slugs in axes.items()
            if axis != "ingredient_entity"
            for slug in slugs
        ],
        "catalog_cleaning": {
            "schema_version": CATALOG_SCHEMA_VERSION,
            "parser_version": PARSER_VERSION,
            "cleaner_version": CLEANER_VERSION,
            "taxonomy_version": TAXONOMY_VERSION,
            "upload_id": upload_id,
            "run_id": run_id,
            "source_record_id": source_row.source_record_id,
            "source_row": source_row.row_number,
            "source_sha256": source_sha256,
            "classification_source": classification["classifier_source"],
            "approval_state": "pending_merchant_review",
            "image_source": "merchant_image_url" if facts["image_url"] else "none",
        },
    }
    return {
        "sku": facts["sku"],
        "merchant_id": facts["merchant_id"],
        "title": facts["title"],
        "description": facts["description"],
        "price_cents": facts["price_cents"],
        "currency": facts["currency"],
        "image_url": facts["image_url"],
        "category": "skincare",
        "attributes": attributes,
        "stock": facts["stock"],
        "rating_avg": facts["rating_avg"],
        "rating_count": facts["rating_count"],
        "rating_source": facts["rating_source"],
    }


def _classification_review_reasons(classification: dict[str, Any]) -> list[str]:
    primary_types = [
        assignment
        for assignment in classification["assignments"]
        if assignment["axis"] == "product_type" and assignment["is_primary"]
    ]
    reasons: list[str] = []
    if len(primary_types) != 1:
        reasons.append("exactly one primary product type is required")
    review_codes = {
        "POTENTIAL_PROMPT_INJECTION",
        "UNSUPPORTED_MEDICAL_CLAIM",
        "UNSUPPORTED_SAFETY_CLAIM",
        "AMBIGUOUS_PRODUCT_TYPE",
        "CONFLICTING_SOURCE_FIELDS",
        "INSUFFICIENT_EVIDENCE",
        "UNKNOWN_INGREDIENT",
    }
    reasons.extend(
        warning["message"]
        for warning in classification["warnings"]
        if warning.get("code") in review_codes or warning.get("severity") == "error"
    )
    return reasons


def _approval_plans(
    *,
    preview_hash: str,
    approval_context: dict[str, Any],
    clean_rows: list[Any],
) -> dict[str, dict[str, Any]]:
    ready_skus = sorted(
        json_load(row["canonical_json"], {})["sku"]
        for row in clean_rows
        if row["status"] == "ready"
    )
    unresolved = sum(row["status"] != "ready" for row in clean_rows)
    base_skus = sorted(approval_context["base_catalog_skus"])
    plans: dict[str, dict[str, Any]] = {}
    for mode in ("replace", "upsert"):
        removal_skus = sorted(set(base_skus) - set(ready_skus)) if mode == "replace" else []
        allowed = bool(ready_skus) and not (mode == "replace" and unresolved)
        token_payload = {
            "preview_hash": preview_hash,
            "base_catalog_hash": approval_context["base_catalog_hash"],
            "mode": mode,
            "publish_skus": ready_skus,
            "removal_skus": removal_skus,
        }
        plans[mode] = {
            "allowed": allowed,
            "blocked_reason": (
                "Resolve or remove every held row before replacing the live catalog."
                if mode == "replace" and unresolved
                else ("No cleaned products are ready to publish." if not ready_skus else None)
            ),
            "approval_token": hashlib.sha256(canonical_json(token_payload).encode()).hexdigest(),
            "publish_count": len(ready_skus),
            "publish_skus": ready_skus,
            "removal_count": len(removal_skus),
            "removal_skus": removal_skus,
            "held_count": unresolved,
        }
    return plans


def _preview_hash(
    *,
    source_sha256: str,
    mappings: dict[str, str],
    taxonomy: dict[str, Any],
    rows: list[dict[str, Any]],
) -> str:
    """Everything a merchant approves when they approve a preview.

    Attaching images rewrites canonical rows, so this is recomputed there too - the hash
    binds the approval tokens, and a preview whose pictures changed is a different preview.
    """
    return hashlib.sha256(
        canonical_json(
            {
                "source_sha256": source_sha256,
                "schema_version": CATALOG_SCHEMA_VERSION,
                "parser_version": PARSER_VERSION,
                "cleaner_version": CLEANER_VERSION,
                "taxonomy_version": TAXONOMY_VERSION,
                "mappings": mappings,
                "taxonomy": taxonomy,
                "rows": [
                    {
                        "source_record_id": row["source_record_id"],
                        "status": row["status"],
                        "canonical": row["canonical"],
                        "issues": row["issues"],
                    }
                    for row in rows
                ],
            }
        ).encode()
    ).hexdigest()


def _resolved_mappings(parsed: ParsedCatalog) -> MappingResolution:
    """Settle the column mapping before anything is staged.

    A tie is a hard stop: guessing which of two columns is the price is exactly the kind of
    decision that has to go back to the merchant, and the template exists so it never arises.
    """
    resolution = resolve_mappings(parsed.candidates)
    if resolution.unresolved:
        first = resolution.unresolved[0]
        raise api_error(
            400,
            "BAD_CATALOG",
            f"Two columns could both be {first['target']}: "
            f"{', '.join(first['candidate_columns'])}. Rename or remove one and upload again, "
            f"or start from the catalog template.",
            unresolved_fields=resolution.unresolved,
        )
    return resolution


async def stage_and_clean_catalog(
    *, merchant: dict[str, Any], parsed: ParsedCatalog
) -> dict[str, Any]:
    mapping = _resolved_mappings(parsed)
    upload_id = new_id("upl")
    run_id = new_id("run")
    now = utc_now()
    with transaction() as connection:
        base_catalog = _catalog_snapshot(connection, merchant["merchant_id"])
        approval_state = {
            "approval_context": {
                "base_catalog_hash": base_catalog["hash"],
                "base_catalog_skus": base_catalog["skus"],
            }
        }
        connection.execute(
            "INSERT INTO catalog_sources(upload_id,merchant_id,filename,source_format,source_sha256,"
            "raw_bytes,byte_count,row_count,source_metadata_json,created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (
                upload_id,
                merchant["merchant_id"],
                parsed.filename,
                parsed.source_format,
                parsed.source_sha256,
                parsed.raw_bytes,
                len(parsed.raw_bytes),
                len(parsed.rows),
                canonical_json(parsed.metadata),
                now,
            ),
        )
        for row in parsed.rows:
            connection.execute(
                "INSERT INTO catalog_source_rows(upload_id,source_record_id,row_number,sheet_name,"
                "raw_row_json,raw_row_sha256,created_at) VALUES (?,?,?,?,?,?,?)",
                (
                    upload_id,
                    row.source_record_id,
                    row.row_number,
                    row.sheet_name,
                    canonical_json(row.values),
                    _row_hash(row.values),
                    now,
                ),
            )
        connection.execute(
            "INSERT INTO catalog_clean_runs(run_id,upload_id,merchant_id,schema_version,parser_version,"
            "cleaner_version,taxonomy_version,prompt_hash,model_name,status,mapping_json,created_at) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                run_id,
                upload_id,
                merchant["merchant_id"],
                CATALOG_SCHEMA_VERSION,
                PARSER_VERSION,
                CLEANER_VERSION,
                TAXONOMY_VERSION,
                PROMPT_HASH,
                settings.openai_model
                if not settings.demo_mode and settings.openai_api_key
                else "deterministic",
                "cleaning",
                canonical_json(mapping.mappings),
                now,
            ),
        )
        connection.execute(
            "UPDATE catalog_clean_runs SET publication_json=?,mapping_report_json=? WHERE run_id=?",
            (canonical_json(approval_state), canonical_json(mapping.report()), run_id),
        )

    records = [
        {
            "source_record_id": row.source_record_id,
            "fields": _descriptive_fields(row, mapping.mappings),
        }
        for row in parsed.rows
    ]
    try:
        classifications, classifier_source = await classify_records(records)
        clean_rows: list[dict[str, Any]] = []
        seen_skus: set[str] = set()
        for row in parsed.rows:
            classification = classifications[row.source_record_id]
            issues: list[str] = []
            try:
                facts = normalize_locked_facts(row, merchant, mapping.mappings)
                if facts["sku"] in seen_skus:
                    raise ProductFactError(
                        [f"duplicate sku {facts['sku']}; each source row must be unique"]
                    )
                seen_skus.add(facts["sku"])
            except ProductFactError as exc:
                facts = None
                issues.extend(exc.issues)
            review_reasons = _classification_review_reasons(classification)
            if facts is None:
                status = "rejected"
                canonical = None
            else:
                canonical = _canonical_product(
                    facts=facts,
                    classification=classification,
                    upload_id=upload_id,
                    run_id=run_id,
                    source_row=row,
                    source_sha256=parsed.source_sha256,
                )
                if review_reasons:
                    status = "review_required"
                    issues.extend(review_reasons)
                else:
                    status = "ready"
            clean_rows.append(
                {
                    "source_record_id": row.source_record_id,
                    "row_number": row.row_number,
                    "status": status,
                    "locked_facts": facts,
                    "classification": classification,
                    "canonical": canonical,
                    "issues": issues,
                    "classifier_source": classification["classifier_source"],
                }
            )

        summary = {
            "input_rows": len(parsed.rows),
            "ready": sum(row["status"] == "ready" for row in clean_rows),
            "review_required": sum(row["status"] == "review_required" for row in clean_rows),
            "rejected": sum(row["status"] == "rejected" for row in clean_rows),
            "fallback_rows": sum("failover" in row["classifier_source"] for row in clean_rows),
        }
        assert summary["input_rows"] == (
            summary["ready"] + summary["review_required"] + summary["rejected"]
        )
        taxonomy = build_taxonomy(
            [row["classification"] for row in clean_rows if row["status"] == "ready"]
        )
        preview_hash = _preview_hash(
            source_sha256=parsed.source_sha256,
            mappings=mapping.mappings,
            taxonomy=taxonomy,
            rows=clean_rows,
        )
        diagnostics = await summarize_upload(
            rows=clean_rows, summary=summary, mapping_report=mapping.report()
        )
        completed_at = utc_now()
        with transaction() as connection:
            for row in clean_rows:
                connection.execute(
                    "INSERT INTO catalog_clean_rows(run_id,upload_id,source_record_id,row_number,status,"
                    "locked_facts_json,classification_json,canonical_json,issues_json,classifier_source,"
                    "created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    (
                        run_id,
                        upload_id,
                        row["source_record_id"],
                        row["row_number"],
                        row["status"],
                        canonical_json(row["locked_facts"])
                        if row["locked_facts"] is not None
                        else None,
                        canonical_json(row["classification"]),
                        canonical_json(row["canonical"]) if row["canonical"] is not None else None,
                        canonical_json(row["issues"]),
                        row["classifier_source"],
                        completed_at,
                    ),
                )
            connection.execute(
                "UPDATE catalog_clean_runs SET classifier_source=?,status='review_ready',taxonomy_json=?,"
                "summary_json=?,diagnostics_json=?,preview_hash=?,completed_at=? WHERE run_id=?",
                (
                    classifier_source,
                    canonical_json(taxonomy),
                    canonical_json(summary),
                    canonical_json(diagnostics),
                    preview_hash,
                    completed_at,
                    run_id,
                ),
            )
    except BaseException:
        with transaction() as connection:
            connection.execute(
                "UPDATE catalog_clean_runs SET status='failed',completed_at=? WHERE run_id=?",
                (utc_now(), run_id),
            )
        raise
    return catalog_upload_preview(merchant["merchant_id"], upload_id)


def catalog_image_url(image_id: str) -> str:
    return f"{settings.catalog_image_base_url.rstrip('/')}/catalog/images/{image_id}"

async def attach_product_images(
    *, merchant: dict[str, Any], content: bytes, filename: str
) -> dict[str, Any]:
    """Bind a ZIP of photos to a merchant's live products, by file name.

    Photos are the one thing merchants add after the fact, so this deliberately does not wait
    for a catalog upload: it matches against the products that are already in the catalog and
    updates them in place. It only ever writes `image_url` - never a price, a title or stock -
    so it cannot change what a product *is*, only how it looks.
    """
    merchant_id = merchant["merchant_id"]
    try:
        images, skipped = extract_image_archive(content, filename)
    except ImageArchiveError as exc:
        raise api_error(400, "BAD_IMAGE_ARCHIVE", str(exc)) from exc

    with connect() as connection:
        rows = connection.execute(
            "SELECT sku,title,image_url FROM products WHERE merchant_id=? ORDER BY sku",
            (merchant_id,),
        ).fetchall()
    if not rows:
        raise api_error(
            409,
            "NO_PRODUCTS",
            "Publish a catalog before adding photos - there is nothing to attach them to yet.",
        )

    products = [dict(row) for row in rows]
    targets = [
        ImageTarget(
            source_record_id=product["sku"],
            row_number=index,
            sku=product["sku"],
            title=product["title"],
        )
        for index, product in enumerate(products, start=1)
    ]
    matches, match_source = await match_images(images, targets)

    now = utc_now()
    stored: list[dict[str, Any]] = []
    matched_skus: set[str] = set()
    with transaction() as connection:
        for image in images:
            match = matches.get(image.entry_name)
            if not match:
                # Nothing references an unmatched photo, so storing its bytes would leave an
                # orphan - and hand the merchant's browser a URL that 404s the moment the
                # sweep below runs.
                stored.append(
                    {
                        "entry_name": image.entry_name,
                        "content_type": image.content_type,
                        "byte_count": image.byte_count,
                        "url": None,
                        "matched": False,
                    }
                )
                continue
            image_id = new_id("img")
            connection.execute(
                "INSERT INTO catalog_images(image_id,merchant_id,archive_name,entry_name,"
                "stem,content_type,byte_count,sha256,image_bytes,created_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?)",
                (
                    image_id,
                    merchant_id,
                    filename,
                    image.entry_name,
                    image.stem,
                    image.content_type,
                    image.byte_count,
                    image.sha256,
                    image.data,
                    now,
                ),
            )
            record = {
                "image_id": image_id,
                "entry_name": image.entry_name,
                "content_type": image.content_type,
                "byte_count": image.byte_count,
                "url": catalog_image_url(image_id),
                "matched": True,
                **match,
            }
            stored.append(record)
            sku = match["source_record_id"]
            matched_skus.add(sku)
            previous = connection.execute(
                "SELECT image_url FROM products WHERE merchant_id=? AND sku=?",
                (merchant_id, sku),
            ).fetchone()
            connection.execute(
                "UPDATE products SET image_url=?,updated_at=? WHERE merchant_id=? AND sku=?",
                (record["url"], now, merchant_id, sku),
            )
            record["replaced_url"] = previous["image_url"] if previous else None
            record["sku"] = sku
            record["title"] = next(p["title"] for p in products if p["sku"] == sku)

        # Photos this archive replaced are now referenced by nothing, and keeping them would
        # grow the database every time a merchant corrects a picture.
        live = connection.execute(
            "SELECT image_url FROM products WHERE merchant_id=? AND image_url IS NOT NULL",
            (merchant_id,),
        ).fetchall()
        in_use = {
            row["image_url"].rsplit("/", 1)[-1]
            for row in live
            if "/catalog/images/" in row["image_url"]
        }
        for row in connection.execute(
            "SELECT image_id FROM catalog_images WHERE merchant_id=?", (merchant_id,)
        ).fetchall():
            if row["image_id"] not in in_use:
                connection.execute(
                    "DELETE FROM catalog_images WHERE image_id=?", (row["image_id"],)
                )

    return {
        "archive": filename,
        "match_source": match_source,
        "image_count": len(images),
        "matched_count": len(matched_skus),
        "product_count": len(products),
        "unmatched_images": [image["entry_name"] for image in stored if not image["matched"]],
        "products_without_images": [
            product["sku"]
            for product in products
            if product["sku"] not in matched_skus and not product["image_url"]
        ],
        "skipped_entries": skipped,
        "images": stored,
    }


def catalog_upload_preview(
    merchant_id: str,
    upload_id: str,
    *,
    offset: int = 0,
    limit: int = MAX_PREVIEW_ROWS,
) -> dict[str, Any]:
    if offset < 0 or not 1 <= limit <= 200:
        raise api_error(400, "BAD_PAGINATION", "Use offset >= 0 and limit between 1 and 200.")
    with connect() as connection:
        run = connection.execute(
            "SELECT r.*,s.filename,s.source_format,s.source_sha256,s.byte_count,s.row_count,"
            "s.source_metadata_json FROM catalog_clean_runs r JOIN catalog_sources s "
            "ON s.upload_id=r.upload_id WHERE r.upload_id=? AND r.merchant_id=?",
            (upload_id, merchant_id),
        ).fetchone()
        if not run:
            raise api_error(404, "NO_CATALOG_UPLOAD", "The catalog upload was not found.")
        rows = connection.execute(
            "SELECT * FROM catalog_clean_rows WHERE run_id=? ORDER BY row_number",
            (run["run_id"],),
        ).fetchall()
    summary = json_load(run["summary_json"], {})
    publication_state = json_load(run["publication_json"], {})
    approval_context = publication_state.get("approval_context", {})
    if not approval_context:
        raise api_error(
            409,
            "CATALOG_APPROVAL_CONTEXT_MISSING",
            "This preview predates safe catalog approvals and must be uploaded again.",
        )
    plans = _approval_plans(
        preview_hash=run["preview_hash"],
        approval_context=approval_context,
        clean_rows=list(rows),
    )
    page_rows = rows[offset : offset + limit]
    previews = [
        {
            "source_record_id": row["source_record_id"],
            "row": row["row_number"],
            "status": row["status"],
            "canonical": json_load(row["canonical_json"], None),
            "classification": json_load(row["classification_json"], {}),
            "issues": json_load(row["issues_json"], []),
            "classifier_source": row["classifier_source"],
        }
        for row in page_rows
    ]
    errors = [
        {
            "row": row["row_number"],
            "reason": "; ".join(json_load(row["issues_json"], [])) or row["status"],
            "status": row["status"],
        }
        for row in page_rows
        if row["status"] != "ready"
    ]
    ready = int(summary.get("ready", 0))
    skipped = int(summary.get("review_required", 0)) + int(summary.get("rejected", 0))
    return {
        "upload_id": upload_id,
        "run_id": run["run_id"],
        "status": run["status"],
        "schema_version": run["schema_version"],
        "parser_version": run["parser_version"],
        "cleaner_version": run["cleaner_version"],
        "taxonomy_version": run["taxonomy_version"],
        "preview_hash": run["preview_hash"],
        "approval_required": run["status"] == "review_ready",
        "ready": ready,
        "ingested": ready if run["status"] == "published" else 0,
        "skipped": skipped,
        "live_products_changed": run["status"] == "published",
        "errors": errors,
        "partial_success": bool(ready and skipped),
        "source": {
            "filename": run["filename"],
            "format": run["source_format"],
            "sha256": run["source_sha256"],
            "byte_count": run["byte_count"],
            "row_count": run["row_count"],
            "metadata": json_load(run["source_metadata_json"], {}),
        },
        "mappings": json_load(run["mapping_json"], {}),
        "mapping_report": json_load(run["mapping_report_json"], {}),
        "diagnostics": json_load(run["diagnostics_json"], {}),
        "summary": summary,
        "taxonomy": json_load(run["taxonomy_json"], {}),
        "approval": {
            "base_catalog_hash": approval_context["base_catalog_hash"],
            "reviewed_row_count_required": len(rows),
            "modes": plans,
        },
        "classifier": {
            "source": run["classifier_source"],
            "model": run["model_name"],
            "prompt_hash": run["prompt_hash"],
        },
        "products": previews,
        "pagination": {
            "offset": offset,
            "limit": limit,
            "returned": len(page_rows),
            "total": len(rows),
            "next_offset": offset + len(page_rows) if offset + len(page_rows) < len(rows) else None,
        },
        "preview_truncated": offset > 0 or offset + len(page_rows) < len(rows),
    }


def approve_catalog_upload(
    *,
    merchant_id: str,
    upload_id: str,
    approval_token: str,
    reviewed_row_count: int,
    mode: str = "replace",
) -> dict[str, Any]:
    if mode not in {"replace", "upsert"}:
        raise api_error(400, "BAD_APPROVAL_MODE", "Approval mode must be replace or upsert.")
    with transaction() as connection:
        run = connection.execute(
            "SELECT * FROM catalog_clean_runs WHERE upload_id=? AND merchant_id=?",
            (upload_id, merchant_id),
        ).fetchone()
        if not run:
            raise api_error(404, "NO_CATALOG_UPLOAD", "The catalog upload was not found.")
        all_rows = connection.execute(
            "SELECT * FROM catalog_clean_rows WHERE run_id=? ORDER BY row_number",
            (run["run_id"],),
        ).fetchall()
        publication_state = json_load(run["publication_json"], {})
        if run["status"] == "published":
            publication = publication_state.get("result", {})
            if (
                run["publish_mode"] != mode
                or publication_state.get("approved_token") != approval_token
            ):
                raise api_error(
                    409,
                    "APPROVAL_REPLAY_MISMATCH",
                    "This catalog was already published with a different approval plan.",
                )
            return {**publication, "idempotent_replay": True}
        if run["status"] != "review_ready":
            raise api_error(409, "CATALOG_NOT_READY", "The catalog is not ready for approval.")
        approval_context = publication_state.get("approval_context", {})
        if not approval_context:
            raise api_error(
                409,
                "CATALOG_APPROVAL_CONTEXT_MISSING",
                "This upload predates safe approvals and must be uploaded again.",
            )
        plans = _approval_plans(
            preview_hash=run["preview_hash"],
            approval_context=approval_context,
            clean_rows=list(all_rows),
        )
        plan = plans[mode]
        if reviewed_row_count != len(all_rows):
            raise api_error(
                409,
                "CATALOG_REVIEW_INCOMPLETE",
                "Review every catalog row before approval.",
                reviewed_row_count=reviewed_row_count,
                required_row_count=len(all_rows),
            )
        if not plan["allowed"]:
            code = (
                "CATALOG_REPLACE_HAS_UNRESOLVED_ROWS"
                if mode == "replace" and plan["held_count"]
                else "NO_READY_PRODUCTS"
            )
            raise api_error(409, code, plan["blocked_reason"])
        if plan["approval_token"] != approval_token:
            raise api_error(
                409,
                "STALE_CATALOG_PREVIEW",
                "The catalog, approval mode, or removal plan changed; review it again.",
            )
        current_catalog = _catalog_snapshot(connection, merchant_id)
        if current_catalog["hash"] != approval_context["base_catalog_hash"]:
            raise api_error(
                409,
                "CATALOG_BASE_CHANGED",
                "The live catalog changed after this preview was created; upload or review it again.",
            )
        rows = [row for row in all_rows if row["status"] == "ready"]
        products = [json_load(row["canonical_json"], {}) for row in rows]
        for product in products:
            conflict = connection.execute(
                "SELECT merchant_id FROM products WHERE sku=? AND merchant_id<>?",
                (product["sku"], merchant_id),
            ).fetchone()
            if conflict:
                raise api_error(
                    409,
                    "SKU_SCOPE_CONFLICT",
                    f"SKU {product['sku']} already belongs to another merchant.",
                )
        now = utc_now()
        removed = 0
        if mode == "replace":
            for sku in plan["removal_skus"]:
                removed += connection.execute(
                    "DELETE FROM products WHERE merchant_id=? AND sku=?",
                    (merchant_id, sku),
                ).rowcount
        for product in products:
            product["attributes"]["catalog_cleaning"]["approval_state"] = (
                "merchant_approved_agent_classification"
            )
            connection.execute(
                "INSERT INTO products(sku,merchant_id,title,description,price_cents,currency,image_url,"
                "category,attributes_json,stock,rating_avg,rating_count,rating_source,created_at,updated_at) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(sku) DO UPDATE SET "
                "title=excluded.title,description=excluded.description,price_cents=excluded.price_cents,"
                "currency=excluded.currency,image_url=excluded.image_url,category=excluded.category,"
                "attributes_json=excluded.attributes_json,stock=excluded.stock,"
                "rating_avg=excluded.rating_avg,rating_count=excluded.rating_count,"
                "rating_source=excluded.rating_source,updated_at=excluded.updated_at",
                (
                    product["sku"],
                    merchant_id,
                    product["title"],
                    product["description"],
                    product["price_cents"],
                    product["currency"],
                    product["image_url"],
                    product["category"],
                    canonical_json(product["attributes"]),
                    product["stock"],
                    product["rating_avg"],
                    product["rating_count"],
                    product["rating_source"],
                    now,
                    now,
                ),
            )
            connection.execute(
                "UPDATE catalog_clean_rows SET canonical_json=? "
                "WHERE run_id=? AND source_record_id=?",
                (
                    canonical_json(product),
                    run["run_id"],
                    product["attributes"]["catalog_cleaning"]["source_record_id"],
                ),
            )
        summary = json_load(run["summary_json"], {})
        publication = {
            "upload_id": upload_id,
            "status": "published",
            "mode": mode,
            "published": len(products),
            "removed": removed,
            "skipped": int(summary.get("review_required", 0)) + int(summary.get("rejected", 0)),
        }
        published_state = {
            "approval_context": approval_context,
            "approved_token": approval_token,
            "result": publication,
        }
        connection.execute(
            "UPDATE catalog_clean_runs SET status='published',publish_mode=?,publication_json=?,"
            "approved_at=? WHERE run_id=?",
            (mode, canonical_json(published_state), now, run["run_id"]),
        )
    return {**publication, "idempotent_replay": False}
