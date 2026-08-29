"""The distributable catalog workbook.

Handing merchants a template is how the importer knows what a column means without asking a
model at upload time: the headers here are exactly the aliases `catalog_mapping` already
recognises, so a file filled in from this template maps deterministically, for free, every
time. Only five columns are required; the rest are evidence the assistant uses when present
and quietly skips when blank.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

TEMPLATE_VERSION = "skincare-catalog-template.v1"
TEMPLATE_FILENAME = "skincare-catalog-template.xlsx"

# (header, required, width, guidance) - header must match a FIELD_ALIASES entry exactly.
COLUMNS: tuple[tuple[str, bool, int, str], ...] = (
    ("sku", True, 16, "Your own product code. Must be unique, and stays the product's identity across uploads."),
    ("title", True, 30, "The product name shoppers see."),
    ("price", True, 12, "Selling price in dollars and cents, for example 29.90. No currency symbols."),
    ("stock", True, 10, "How many you have. A whole number, 0 or more."),
    ("ingredients", True, 42, "Ingredient or INCI list, comma separated. This is what recommendations are grounded in."),
    ("description", False, 46, "Your own product copy. The more factual detail, the better the assistant answers."),
    ("product_type", False, 18, "Your own wording, e.g. cleanser, serum, sunscreen."),
    ("skin_types", False, 24, "Who it suits, separated by | - e.g. dry|sensitive. Only state what you can stand behind."),
    ("concerns", False, 24, "What it helps with, separated by | - e.g. dryness|redness."),
    ("usage_time", False, 14, "morning, evening, or both."),
    ("texture", False, 14, "e.g. gel, cream, balm, oil."),
    ("fragrance_free", False, 14, "yes or no. Leave blank if you are not certain."),
    ("excludes", False, 26, "Free-from claims you can stand behind, e.g. alcohol|essential oils."),
    ("size_ml", False, 10, "Volume in millilitres, a positive number."),
    ("image_url", False, 34, "An https:// link to the photo. Leave blank to upload photos as a ZIP instead."),
    ("rating_avg", False, 11, "Average rating from 0 to 5. Leave blank if you have none."),
    ("rating_count", False, 12, "How many ratings that average is based on. Required if you fill in rating_avg."),
    ("currency", False, 10, "Only if it differs from your store currency - it must match, nothing is converted."),
)

EXAMPLE_ROWS: tuple[tuple[object, ...], ...] = (
    (
        "MYSA-CLN-100", "Gentle Cloud Cleanser", 29.90, 24,
        "aqua, glycerin, panthenol, sodium cocoyl glycinate",
        "A gentle hydrating face wash that cleans without stripping.",
        "face wash", "dry|sensitive", "dryness", "morning", "gel", "yes",
        "alcohol|essential oils", 150, "", 4.6, 128, "",
    ),
    (
        "MYSA-SER-030", "Calm Gel Serum", 35.00, 12,
        "aqua, niacinamide, panthenol",
        "A light serum for visible redness and oiliness.",
        "serum", "oily|combination", "redness|oiliness", "evening", "gel", "no",
        "", 30, "", "", "", "",
    ),
)

HEADER_FILL = PatternFill("solid", fgColor="E8EFE4")
REQUIRED_FILL = PatternFill("solid", fgColor="D6E4CF")
EXAMPLE_FONT = Font(italic=True, color="7A8578")


def build_template() -> bytes:
    """Build the workbook fresh on each request; it is small and never cached stale."""
    book = Workbook()

    sheet = book.active
    sheet.title = "Products"
    sheet.append([column for column, _, _, _ in COLUMNS])
    for index, (_, required, width, guidance) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = Font(bold=True)
        cell.fill = REQUIRED_FILL if required else HEADER_FILL
        cell.alignment = Alignment(vertical="center")
        cell.comment = None
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"

    for row in EXAMPLE_ROWS:
        sheet.append(list(row))
    for row_index in range(2, 2 + len(EXAMPLE_ROWS)):
        for column_index in range(1, len(COLUMNS) + 1):
            sheet.cell(row=row_index, column=column_index).font = EXAMPLE_FONT

    # yes/no is the only value list worth enforcing in the sheet; everything else is prose
    # the importer validates on upload, where the message can actually explain itself.
    fragrance_column = get_column_letter([c for c, _, _, _ in COLUMNS].index("fragrance_free") + 1)
    validation = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"{fragrance_column}2:{fragrance_column}1000")

    guide = book.create_sheet("How to fill this in")
    guide.append(["Column", "Required?", "What to put"])
    for cell in guide[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    for column, required, _, guidance in COLUMNS:
        guide.append([column, "Required" if required else "Optional", guidance])
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 12
    guide.column_dimensions["C"].width = 92
    for row in guide.iter_rows(min_row=2):
        row[2].alignment = Alignment(wrap_text=True, vertical="top")
    guide.append([])
    for note in (
        "Delete the two grey example rows before you upload.",
        "Keep the header row exactly as it is - that is how each column is recognised.",
        "Leave a cell blank when you do not know. A blank is honest; a guess is not.",
        "Do not write instructions to the assistant in any cell. Product copy only.",
        "Paste values, not formulas: a cell holding a formula is held back for review.",
        "Photos: either fill in image_url, or upload a ZIP of photos named after each product.",
    ):
        guide.append(["", "", note])

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()
